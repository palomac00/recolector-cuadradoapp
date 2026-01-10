#!/usr/bin/env python3
import sys
import json
import pandas as pd
import os
from datetime import datetime, timedelta
import pytz
from pathlib import Path
from openpyxl.styles import Font

TZ_AR = pytz.timezone('America/Argentina/Buenos_Aires')

def get_fecha_excel():
    """Nombre del Excel de HOY: horarios-141-YYYY-MM-DD.xlsx"""
    return f"data/horarios-141-{datetime.now(TZ_AR).strftime('%Y-%m-%d')}.xlsx"

def parse_arrivals(json_file):
    """Parse Cuadrado API JSON → lista de arrivals"""
    with open(json_file, 'r') as f:
        data = json.load(f)
    
    arrivals = []
    now = datetime.now(TZ_AR)
    hora_scraping = now.strftime("%H:%M:%S")
    
    for arrival in data['arribos']:
        minutos = arrival['tiempo']
        bandera = arrival['bandera']
        parada = arrival.get('parada', 'LP1912')  # Ahora viene del merge.py
        
        eta = now + timedelta(minutes=minutos)
        hora_eta = eta.strftime("%H:%M")
        
        arrivals.append({
            'Hora_Scrap': hora_scraping,
            'Hora_Llegada': hora_eta,
            'Linea': bandera,
            'Minutos': minutos,
            'Parada': parada
        })
    
    return arrivals

def cargar_excel_dia():
    """Carga el Excel de HOY o retorna DataFrames vacíos"""
    archivo_hoy = get_fecha_excel()
    Path("data").mkdir(exist_ok=True)
    
    if not os.path.exists(archivo_hoy):
        return {
            'LP1912': pd.DataFrame(),
            'LP1912-215': pd.DataFrame(),
            '6203-6173': pd.DataFrame()
        }
    
    try:
        excel_file = pd.ExcelFile(archivo_hoy)
        datos = {}
        
        for sheet in ['LP1912', 'LP1912-215', '6203-6173']:
            if sheet in excel_file.sheet_names:
                df = pd.read_excel(archivo_hoy, sheet_name=sheet, skiprows=4)
                columnas_validas = ['Hora_Scrap', 'Hora_Llegada', 'Linea', 'Minutos', 'Parada']
                df = df[[col for col in columnas_validas if col in df.columns]]
                df = df.dropna(how='all')
                
                # Filtrar por parada según el sheet
                if sheet == 'LP1912':
                    df = df[df['Parada'] == 'LP1912']
                elif sheet == 'LP1912-215':
                    df = df[df['Parada'] == 'LP1912']
                elif sheet == '6203-6173':
                    df = df[df['Parada'].isin(['L6173', 'L6203'])]
                
                datos[sheet] = df
            else:
                datos[sheet] = pd.DataFrame()
        
        return datos
    except Exception as e:
        print(f"⚠️ Error cargando {archivo_hoy}: {e}")
        return {
            'LP1912': pd.DataFrame(),
            'LP1912-215': pd.DataFrame(),
            '6203-6173': pd.DataFrame()
        }

def guardar_excel_dia(horarios_nuevos):
    """Actualiza SOLO el Excel de HOY - SIN DUPLICADOS"""
    datos_existentes = cargar_excel_dia()
    ahora = datetime.now(TZ_AR)
    archivo_hoy = get_fecha_excel()
    Path("data").mkdir(exist_ok=True)
    
    # LP1912 - solo los de parada LP1912
    df_nuevos_lp = pd.DataFrame([h for h in horarios_nuevos if h['Parada'] == 'LP1912'])
    if not datos_existentes['LP1912'].empty:
        df_lp1912 = pd.concat([datos_existentes['LP1912'], df_nuevos_lp], ignore_index=True)
        df_lp1912 = df_lp1912.drop_duplicates(subset=['Hora_Llegada', 'Linea']).reset_index(drop=True)
    else:
        df_lp1912 = df_nuevos_lp
    
    df_lp1912 = df_lp1912.sort_values('Hora_Llegada').reset_index(drop=True)
    
    # LP1912-215 (solo línea 215 de LP1912)
    nuevos_215 = [h for h in horarios_nuevos if h['Parada'] == 'LP1912' and '215' in h.get('Linea', '')]
    if nuevos_215:
        df_nuevos_215 = pd.DataFrame(nuevos_215)
        if not datos_existentes['LP1912-215'].empty:
            df_215 = pd.concat([datos_existentes['LP1912-215'], df_nuevos_215], ignore_index=True)
            df_215 = df_215.drop_duplicates(subset=['Hora_Llegada', 'Linea']).reset_index(drop=True)
        else:
            df_215 = df_nuevos_215
    else:
        df_215 = datos_existentes['LP1912-215']
    
    df_215 = df_215.sort_values('Hora_Llegada').reset_index(drop=True)
    
    # 6203-6173 (L6203 + L6173)
    df_nuevos_comb = pd.DataFrame([h for h in horarios_nuevos if h['Parada'] in ['L6173', 'L6203']])
    if not datos_existentes['6203-6173'].empty:
        df_6203_6173 = pd.concat([datos_existentes['6203-6173'], df_nuevos_comb], ignore_index=True)
        df_6203_6173 = df_6203_6173.drop_duplicates(subset=['Hora_Llegada', 'Linea', 'Parada']).reset_index(drop=True)
    else:
        df_6203_6173 = df_nuevos_comb
    
    df_6203_6173 = df_6203_6173.sort_values('Hora_Llegada').reset_index(drop=True)
    
    with pd.ExcelWriter(archivo_hoy, engine='openpyxl') as writer:
        df_lp1912.to_excel(writer, sheet_name='LP1912', index=False, startrow=4)
        df_215.to_excel(writer, sheet_name='LP1912-215', index=False, startrow=4)
        df_6203_6173.to_excel(writer, sheet_name='6203-6173', index=False, startrow=4)
        
        sheets_info = {
            'LP1912': (df_lp1912, 'LP1912'),
            'LP1912-215': (df_215, 'LP1912-215'),
            '6203-6173': (df_6203_6173, '6203-6173')
        }
        
        for sheet_name, (df, titulo) in sheets_info.items():
            ws = writer.sheets[sheet_name]
            ws['A1'] = f'LÍNEA 141 - {titulo} - {ahora.strftime("%d/%m/%Y")}'
            ws['A1'].font = Font(bold=True)
            ws['A2'] = f'Última actualización: {ahora.strftime("%H:%M:%S")}'
            ws['A2'].font = Font(bold=True)
            ws['A3'] = f'Total filas: {len(df)}'
            ws['A3'].font = Font(bold=True)

    print(f"💾 Excel actualizado: {archivo_hoy}")
    print(f"   LP1912: {len(df_lp1912)} filas")
    print(f"   LP1912-215: {len(df_215)} filas")
    print(f"   6203-6173: {len(df_6203_6173)} filas")

def main():
    if len(sys.argv) != 2:
        print("Usage: python parse_json.py input.json")
        sys.exit(1)
    
    json_file = sys.argv[1]
    arrivals = parse_arrivals(json_file)
    
    print(f"📊 {len(arrivals)} horarios parseados")
    print(f"🚐 {len([a for a in arrivals if '215' in a['Linea']])} buses 215 encontrados")
    
    guardar_excel_dia(arrivals)
    
    print("🎉 Recolector COMPLETO")

if __name__ == "__main__":
    main()
